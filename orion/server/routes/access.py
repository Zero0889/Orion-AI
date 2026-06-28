"""server.routes.access — Control de acceso por huella.

Endpoints
---------

ESP32 (firmware) → Orion:
  POST   /api/access/event              registra un evento del sensor

Frontend (panel "Acceso"):
  GET    /api/access/users              lista de personas enroladas
  POST   /api/access/users              crea un mapping huella→nombre
  PATCH  /api/access/users/{id}         actualiza nombre/teléfono/activo
  DELETE /api/access/users/{id}         elimina mapping (no toca eventos)

  GET    /api/access/events?limit=…     lista cruda paginada
  GET    /api/access/events/count       count total

  GET    /api/access/daily?since=YYYY-MM-DD
                                        reporte agrupado (la "tabla excel")
  GET    /api/access/export.csv         export CSV del reporte diario
  GET    /api/access/export.xlsx        export XLSX del reporte diario

Tras cada POST event:
  · publica `access.event` en el bus (web invalida sus queries en tiempo real)
  · si Telegram está configurado y el evento es GRANTED/DENIED, manda mensaje
"""

from __future__ import annotations

import csv
import io
from typing import Literal

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import Response
from pydantic import BaseModel, Field

from orion.adapters.iot import access_control as ac
from orion.adapters.messaging.telegram import TelegramClient, load_telegram_config
from orion.core.logger import get_logger

log = get_logger("server.access")

router = APIRouter()


# ── Pydantic models ─────────────────────────────────────────────────────


class UserCreate(BaseModel):
    fingerprint_id: int = Field(..., ge=0, le=127)
    name: str = Field(..., min_length=1, max_length=80)
    phone: str = Field(default="", max_length=40)
    active: bool = True


class UserUpdate(BaseModel):
    name: str | None = Field(default=None, max_length=80)
    phone: str | None = Field(default=None, max_length=40)
    active: bool | None = None


class EventIn(BaseModel):
    """Payload que manda el ESP32 cada vez que el sensor lee una huella."""

    fingerprint_id: int = Field(..., ge=-1, le=127)
    event_type: Literal["GRANTED", "DENIED", "ENROLLED"] = "GRANTED"
    esp_id: str = Field(default="", max_length=40)
    confidence: int = Field(default=0, ge=0, le=10_000)


# ── Response models (para que aparezcan en el OpenAPI y se generen
#    como Schemas[] en el frontend) ──────────────────────────────────────


class AccessUserOut(BaseModel):
    id: str
    fingerprint_id: int
    name: str
    phone: str
    active: bool
    created: str


class AccessEventOut(BaseModel):
    id: str
    fingerprint_id: int  # -1 si el sensor no reconoció
    event_type: Literal["GRANTED", "DENIED", "ENROLLED"]
    esp_id: str
    confidence: int
    timestamp: str
    user_name: str | None


class AccessEventsPageOut(BaseModel):
    items: list[AccessEventOut]
    total: int
    limit: int
    offset: int


class AccessEventsCountOut(BaseModel):
    count: int


class AccessDailyRowOut(BaseModel):
    fingerprint_id: int
    name: str
    fecha: str  # YYYY-MM-DD
    entrada: str  # HH:MM
    salida: str  # HH:MM
    tiempo_minutos: int
    tiempo_legible: str  # "8 h 57 min"
    eventos_dia: int


# ── Users ────────────────────────────────────────────────────────────────


@router.get("/users", response_model=list[AccessUserOut])
def list_users():
    return [u.to_dict() for u in ac.list_users()]


@router.post("/users", status_code=201, response_model=AccessUserOut)
def create_user(body: UserCreate, request: Request):
    try:
        user = ac.add_user(
            fingerprint_id=body.fingerprint_id,
            name=body.name,
            phone=body.phone,
            active=body.active,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    _publish(request, "access.user_changed", {"id": user.id})
    return user.to_dict()


@router.patch("/users/{user_id}", response_model=AccessUserOut)
def patch_user(user_id: str, body: UserUpdate, request: Request):
    try:
        user = ac.update_user(
            user_id,
            name=body.name,
            phone=body.phone,
            active=body.active,
        )
    except ValueError as e:
        # Distingue "no existe" (404) de "valor inválido" (400)
        msg = str(e)
        if "no existe" in msg.lower():
            raise HTTPException(status_code=404, detail=msg) from e
        raise HTTPException(status_code=400, detail=msg) from e
    _publish(request, "access.user_changed", {"id": user.id})
    return user.to_dict()


@router.delete("/users/{user_id}", status_code=204)
def remove_user(user_id: str, request: Request) -> None:
    ok = ac.delete_user(user_id)
    if not ok:
        raise HTTPException(status_code=404, detail=f"Usuario {user_id} no existe.")
    _publish(request, "access.user_changed", {"id": user_id, "deleted": True})


# ── Events ───────────────────────────────────────────────────────────────


@router.post("/event", status_code=201, response_model=AccessEventOut)
def post_event(body: EventIn, request: Request):
    """Endpoint que el ESP32 invoca después de cada lectura del AS608."""
    try:
        ev = ac.record_event(
            fingerprint_id=body.fingerprint_id,
            event_type=body.event_type,
            esp_id=body.esp_id,
            confidence=body.confidence,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    _publish(request, "access.event", ev.to_dict())
    _maybe_notify_telegram(ev)
    return ev.to_dict()


@router.get("/events", response_model=AccessEventsPageOut)
def get_events(
    limit: int = Query(default=100, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    fingerprint_id: int | None = Query(default=None, ge=-1, le=127),
    since: str | None = Query(default=None, description="ISO timestamp; >= since"),
    event_type: str | None = Query(default=None),
):
    et: ac.EventType | None = None
    if event_type:
        if event_type not in ac.VALID_EVENT_TYPES:
            raise HTTPException(status_code=400, detail=f"event_type inválido: {event_type}")
        et = event_type  # type: ignore[assignment]
    items = ac.list_events(
        limit=limit, offset=offset, fingerprint_id=fingerprint_id, since=since, event_type=et
    )
    total = ac.count_events(fingerprint_id=fingerprint_id, since=since, event_type=et)
    return {"items": [e.to_dict() for e in items], "total": total, "limit": limit, "offset": offset}


@router.get("/events/count", response_model=AccessEventsCountOut)
def events_count():
    return {"count": ac.count_events()}


# ── Reporte diario (la "tabla excel") ────────────────────────────────────


@router.get("/daily", response_model=list[AccessDailyRowOut])
def daily(
    since: str | None = Query(default=None, description="YYYY-MM-DD; >= since"),
    fingerprint_id: int | None = Query(default=None, ge=0, le=127),
):
    return [r.to_dict() for r in ac.daily_report(since=since, fingerprint_id=fingerprint_id)]


@router.get("/export.csv")
def export_csv(since: str | None = Query(default=None)) -> Response:
    rows = ac.daily_report(since=since)
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(["ID", "Nombre", "Fecha", "Entrada", "Salida", "Tiempo", "Eventos"])
    for i, r in enumerate(rows, start=1):
        w.writerow(
            [
                f"{i:03d}",
                r.name,
                _format_fecha(r.fecha),
                r.entrada,
                r.salida,
                r.tiempo_legible,
                r.eventos_dia,
            ]
        )
    return Response(
        content=buf.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="orion-acceso-diario.csv"'},
    )


@router.get("/export.xlsx")
def export_xlsx(since: str | None = Query(default=None)) -> Response:
    """Exporta la tabla diaria como XLSX con dos hojas:
    - "Reporte diario": una fila por usuario/día (la tabla que pidió el user)
    - "Eventos crudos": todos los registros con confidencia/esp_id/tipo

    Requiere ``openpyxl`` (lo agregamos a pyproject como dep opcional —
    si no está instalado devolvemos 503 con instrucción clara).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise HTTPException(
            status_code=503,
            detail=(
                "Export XLSX requiere openpyxl. Instalalo con: pip install openpyxl "
                "(o usá /api/access/export.csv que no tiene esa dep)."
            ),
        ) from e

    daily_rows = ac.daily_report(since=since)
    event_rows = ac.list_events(limit=1000, since=since)

    wb = Workbook()

    # ── Hoja 1: Reporte diario ──────────────────────────────────────────
    ws_daily = wb.active
    assert ws_daily is not None
    ws_daily.title = "Reporte diario"
    headers_daily = ["ID", "Nombre", "Fecha", "Entrada", "Salida", "Tiempo", "Eventos"]
    ws_daily.append(headers_daily)

    header_font = Font(bold=True, color="FFFFFFFF", name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor="FF1F2937")  # gris oscuro
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws_daily[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, r in enumerate(daily_rows, start=1):
        ws_daily.append(
            [
                f"{i:03d}",
                r.name,
                _format_fecha(r.fecha),
                r.entrada,
                r.salida,
                r.tiempo_legible,
                r.eventos_dia,
            ]
        )

    # Anchos razonables (no auto-size — openpyxl no lo soporta sin medir
    # los strings; pongo valores fijos que cubren el caso común).
    widths_daily = [8, 28, 14, 12, 12, 14, 10]
    for col_idx, width in enumerate(widths_daily, start=1):
        ws_daily.column_dimensions[chr(ord("A") + col_idx - 1)].width = width

    # ── Hoja 2: Eventos crudos ──────────────────────────────────────────
    ws_ev = wb.create_sheet("Eventos crudos")
    headers_ev = ["Timestamp", "Huella ID", "Usuario", "Tipo", "ESP", "Confianza"]
    ws_ev.append(headers_ev)
    for cell in ws_ev[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for ev in event_rows:
        ws_ev.append(
            [
                ev.timestamp,
                ev.fingerprint_id if ev.fingerprint_id >= 0 else "—",
                ev.user_name or "—",
                ev.event_type,
                ev.esp_id or "—",
                ev.confidence,
            ]
        )

    widths_ev = [22, 12, 28, 12, 16, 12]
    for col_idx, width in enumerate(widths_ev, start=1):
        ws_ev.column_dimensions[chr(ord("A") + col_idx - 1)].width = width

    # Freezar el header en ambas hojas para que el scroll lo deje fijo.
    ws_daily.freeze_panes = "A2"
    ws_ev.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="orion-acceso.xlsx"'},
    )


# ── Helpers internos ────────────────────────────────────────────────────


def _publish(request: Request, event: str, payload: dict) -> None:
    bus = getattr(request.app.state, "bus", None)
    if bus is None:
        return
    try:
        bus.publish(event, payload)
    except Exception as e:  # pragma: no cover
        log.warning("publish %s falló: %s", event, e)


def _maybe_notify_telegram(ev: ac.AccessEvent) -> None:
    """Si Telegram está habilitado, manda un mensaje por cada GRANTED o DENIED.
    ENROLLED (registro de huella nueva) no manda — es admin operation.

    Routing del destino:
      1. Si hay supergrupo con topic ``access`` mapeado → al topic.
      2. Si no, fallback al ``default_chat_id`` (chat privado del usuario).
    """
    if ev.event_type not in ("GRANTED", "DENIED"):
        return
    cfg = load_telegram_config()
    if not (cfg.is_configured and cfg.enabled):
        return
    nombre = (
        ev.user_name or f"Huella #{ev.fingerprint_id}" if ev.fingerprint_id >= 0 else "Desconocido"
    )
    hora = ev.timestamp[11:16] if len(ev.timestamp) >= 16 else ev.timestamp
    if ev.event_type == "GRANTED":
        icon, accion = "🔓", "ingresó al domicilio"
    else:
        icon, accion = "⛔", "intentó ingresar y fue denegado"
    location = f" en *{ev.esp_id}*" if ev.esp_id else ""
    msg = f"{icon} *{nombre}* {accion}{location}\n🕒 {hora}"

    target = cfg.resolve_topic("access")
    if target is not None:
        chat_id, thread_id = target
    elif cfg.default_chat_id:
        chat_id, thread_id = cfg.default_chat_id, None
    else:
        return

    try:
        client = TelegramClient(cfg.bot_token)
        client.send_message(chat_id, msg, message_thread_id=thread_id)
    except Exception as e:  # pragma: no cover
        log.warning("notify telegram falló: %s", e)


def _format_fecha(iso: str) -> str:
    """`2026-06-27` → `27/06/2026` (formato latam más legible en planillas)."""
    if len(iso) != 10:
        return iso
    y, m, d = iso.split("-")
    return f"{d}/{m}/{y}"
