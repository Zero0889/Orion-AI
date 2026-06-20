"""
actions.iot.sensor_log — Datalogger persistente de sensores IoT
================================================================
Acumula cada lectura por dispositivo en un buffer en memoria, calcula
el promedio por minuto, y escribe una fila por dispositivo cada 60s
en ``memory/iot_sensor_log.csv``.

El callback de los transports (MQTT / Serial) llama a :func:`record`
cada vez que llega una lectura. El thread interno de flush se
encarga del resto.

Provee :func:`read_csv_bytes` y :func:`read_xlsx_bytes` para que los
endpoints de descarga sirvan el archivo en CSV o Excel formateado.
"""

from __future__ import annotations

import csv
import io
import threading
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from core.logger import get_logger
import contextlib

log = get_logger("iot.sensor_log")


# ── Estado interno ──────────────────────────────────────────────────────
_buffers: dict[str, list[float]] = defaultdict(list)
_lock = threading.Lock()
_log_path: Path | None = None
_flusher_thread: threading.Thread | None = None
_flusher_stop = threading.Event()

FLUSH_INTERVAL_S = 60

# Unidades por tipo de sensor (mismo set que el frontend en sensorPersonality.ts)
UNIT_MAP = {
    "temperature": "°C",
    "humidity": "%",
    "light": "lx",
    "geo": "",
    "count": "",
}

CSV_HEADER = ["timestamp", "device", "value", "unit", "samples"]


# ── Helpers ────────────────────────────────────────────────────────────
def _resolve_path() -> Path:
    global _log_path
    if _log_path is None:
        from config import IOT_CONFIG_PATH

        # IOT_CONFIG_PATH = .../config/iot_config.json → root del proyecto
        root = IOT_CONFIG_PATH.parent.parent
        _log_path = root / "memory" / "iot_sensor_log.csv"
    _log_path.parent.mkdir(parents=True, exist_ok=True)
    return _log_path


def _resolve_unit(device_id: str) -> str:
    """Mira el config del IoT system para sacar la unidad."""
    try:
        from actions.iot import get_system

        sys = get_system()
        dev = sys.cfg.devices.get(device_id)
        if dev and dev.capabilities.sensor:
            return UNIT_MAP.get(dev.capabilities.sensor.lower(), "")
    except Exception:
        pass
    return ""


# ── API pública ────────────────────────────────────────────────────────
def record(device_id: str, raw_value: str) -> None:
    """Acumula una lectura. Llamado desde el callback del transport.

    Reglas para evitar 'valores fantasma' en el log/Sheet:
      * No numéricos (lat/lon, strings) se descartan.
      * NaN / ±Inf se descartan (a veces los sensores devuelven NaN
        cuando se desconectan momentáneamente).
      * Si en un minuto no llega ninguna lectura para un device, no se
        escribe fila para él en :func:`_flush_once`.
    """
    import math

    try:
        v = float(raw_value)
    except (TypeError, ValueError):
        return
    if math.isnan(v) or math.isinf(v):
        return
    with _lock:
        _buffers[device_id].append(v)


def _flush_once() -> int:
    """Vacía el buffer al disco. Devuelve cantidad de filas escritas."""
    with _lock:
        if not _buffers:
            return 0
        snapshot = {k: list(v) for k, v in _buffers.items() if v}
        _buffers.clear()
    if not snapshot:
        return 0

    ts = datetime.now().replace(second=0, microsecond=0).isoformat(sep=" ")
    rows: list[tuple] = []
    for dev_id, values in snapshot.items():
        avg = sum(values) / len(values)
        rows.append((ts, dev_id, round(avg, 3), _resolve_unit(dev_id), len(values)))

    path = _resolve_path()
    fresh = not path.exists() or path.stat().st_size == 0
    with path.open("a", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        if fresh:
            w.writerow(CSV_HEADER)
        for r in rows:
            w.writerow(r)
    return len(rows)


def _flusher_loop() -> None:
    while not _flusher_stop.wait(FLUSH_INTERVAL_S):
        try:
            n = _flush_once()
            if n:
                log.debug("flush: %d filas escritas", n)
        except Exception as e:
            log.warning("flush error: %s", e)


def start() -> None:
    """Arranca el thread de flush periódico. Idempotente."""
    global _flusher_thread
    if _flusher_thread and _flusher_thread.is_alive():
        return
    _resolve_path()
    _flusher_stop.clear()
    _flusher_thread = threading.Thread(
        target=_flusher_loop,
        daemon=True,
        name="iot-sensor-log-flusher",
    )
    _flusher_thread.start()
    log.info("datalogger iniciado (flush cada %ds)", FLUSH_INTERVAL_S)


def stop() -> None:
    """Para el thread y persiste el buffer remanente."""
    _flusher_stop.set()
    with contextlib.suppress(Exception):
        _flush_once()


# ── Lectura para descarga ──────────────────────────────────────────────
def read_csv_bytes() -> bytes:
    """Devuelve el CSV completo. Vacío con header si no hay data aún."""
    # Asegurar que el buffer activo esté en disco antes de exportar
    with contextlib.suppress(Exception):
        _flush_once()
    path = _resolve_path()
    if not path.exists() or path.stat().st_size == 0:
        return (",".join(CSV_HEADER) + "\n").encode("utf-8")
    return path.read_bytes()


def read_xlsx_bytes() -> bytes:
    """Devuelve un XLSX con una hoja 'all' + una hoja por dispositivo."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with contextlib.suppress(Exception):
        _flush_once()

    path = _resolve_path()
    rows: list[list[str]] = []
    if path.exists() and path.stat().st_size > 0:
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            with contextlib.suppress(StopIteration):
                next(reader)  # skip header
            rows = [r for r in reader if r]

    head_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    head_font = Font(color="FFFFFF", bold=True)
    head_align = Alignment(horizontal="center")

    def write_sheet(ws, header: list[str], data_rows: list[list[str]]) -> None:
        ws.append(header)
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = head_align
        for row in data_rows:
            ws.append(row)
        widths = [len(h) for h in header]
        for row in data_rows:
            for i, v in enumerate(row[: len(widths)]):
                widths[i] = max(widths[i], len(str(v)))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 32)
        ws.freeze_panes = "A2"

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "all"
    write_sheet(ws_all, CSV_HEADER, rows)

    # Hojas por dispositivo (sin columna "device", para que sea más limpio)
    by_device: dict[str, list[list[str]]] = defaultdict(list)
    for r in rows:
        if len(r) >= 2:
            by_device[r[1]].append(r)
    for dev_id in sorted(by_device.keys()):
        ws = wb.create_sheet(title=_safe_sheet_name(dev_id))
        dev_rows = [
            [
                r[0],
                r[2] if len(r) > 2 else "",
                r[3] if len(r) > 3 else "",
                r[4] if len(r) > 4 else "",
            ]
            for r in by_device[dev_id]
        ]
        write_sheet(ws, ["timestamp", "value", "unit", "samples"], dev_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe_sheet_name(name: str) -> str:
    """Excel limita: max 31 chars, sin / \\ ? * [ ]."""
    bad = set("\\/?*[]:")
    safe = "".join(c for c in name if c not in bad).strip()[:31]
    return safe or "sheet"
